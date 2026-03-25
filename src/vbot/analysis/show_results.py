# src/vbot/analysis/show_results.py
# vbot — Ergebnisanzeige und Portfolio-Analyse

import os
import sys
import json
import logging
import argparse
from datetime import date
from typing import Optional

import pandas as pd

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

logging.basicConfig(level=logging.WARNING, format='%(levelname)s %(message)s')
logger = logging.getLogger(__name__)

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'vbot', 'strategy', 'configs')
RESULTS_DIR   = os.path.join(PROJECT_ROOT, 'artifacts', 'results')
SETTINGS_FILE = os.path.join(PROJECT_ROOT, 'settings.json')

GREEN  = '\033[0;32m'
YELLOW = '\033[1;33m'
RED    = '\033[0;31m'
CYAN   = '\033[0;36m'
BOLD   = '\033[1m'
NC     = '\033[0m'


# ---------------------------------------------------------------------------
# Modus 1: Einzel-Analyse — alle Configs isoliert testen
# ---------------------------------------------------------------------------

def run_all_configs_isolated(date_from: str, date_to: str, capital: float,
                               configs_filter: list = None):
    from vbot.analysis.backtester import run_backtest, load_ohlcv

    if not os.path.isdir(CONFIGS_DIR):
        print(f"{RED}Kein Configs-Verzeichnis: {CONFIGS_DIR}{NC}")
        return

    cfg_files = sorted(f for f in os.listdir(CONFIGS_DIR)
                       if f.startswith('config_') and f.endswith('.json'))
    if configs_filter:
        cfg_files = [f for f in cfg_files if f in configs_filter]
    if not cfg_files:
        print(f"{YELLOW}Keine Configs gefunden. Erst run_pipeline.sh ausfuehren.{NC}")
        return

    print(f"\n--- vbot Ergebnis-Analyse (Einzel-Modus) ---")
    print(f"Zeitraum: {date_from} bis {date_to} | Startkapital: {capital:.0f} USDT\n")

    results = []
    for fname in cfg_files:
        cfg_path = os.path.join(CONFIGS_DIR, fname)
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception:
            continue

        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue

        print(f"Analysiere: {fname}...")
        df = load_ohlcv(symbol, timeframe, date_from, date_to)
        if df.empty or len(df) < 10:
            print(f"  {YELLOW}Keine Daten — uebersprungen.{NC}")
            continue

        result = run_backtest(df, config, capital, symbol, timeframe)
        fibo_lvl = config.get('signal', {}).get('fibo_tp_level', '?')
        results.append({
            'filename':  fname,
            'symbol':    symbol,
            'timeframe': timeframe,
            'trades':    result.total_trades,
            'win_rate':  result.win_rate,
            'pnl_pct':   result.pnl_pct,
            'max_dd':    result.max_drawdown_pct,
            'avg_rr':    result.avg_rr,
            'end_cap':   result.end_capital,
            'fibo_lvl':  fibo_lvl,
        })

    if not results:
        print(f"{RED}Kein Backtest erfolgreich.{NC}")
        return

    W = 112
    print(f"\n{'='*W}")
    print(f"{'Zusammenfassung aller Einzelstrategien':^{W}}")
    print(f"{'='*W}")
    print(f"  {'Strategie':<24}  {'Trades':>6}  {'WR %':>6}  {'PnL %':>7}  {'MaxDD %':>7}  {'R:R':>5}  {'Fibo':>5}  {'Endkapital':>12}")
    for r in sorted(results, key=lambda x: x['pnl_pct'], reverse=True):
        strat  = f"{r['symbol'].split('/')[0]}/{r['timeframe']}"
        color  = GREEN if r['pnl_pct'] >= 0 else RED
        dd_col = GREEN if r['max_dd'] <= 30 else RED
        print(f"  {strat:<24}  {r['trades']:>6}  {r['win_rate']:>6.1f}  "
              f"{color}{r['pnl_pct']:>7.2f}{NC}  "
              f"{dd_col}{r['max_dd']:>7.2f}{NC}  "
              f"{r['avg_rr']:>5.2f}  {str(r['fibo_lvl']):>5}  "
              f"{color}{r['end_cap']:>11.2f}{NC} USDT")
    print(f"{'='*W}")


# ---------------------------------------------------------------------------
# Modus 2: Manuelle Portfolio-Simulation
# ---------------------------------------------------------------------------

def run_manual_portfolio(date_from: str, date_to: str, capital: float,
                          selected_files: list):
    from vbot.analysis.backtester import run_backtest, load_ohlcv
    from vbot.analysis.portfolio_simulator import run_portfolio_simulation

    strategies_data = {}
    for fname in selected_files:
        cfg_path = os.path.join(CONFIGS_DIR, fname)
        if not os.path.exists(cfg_path):
            print(f"{YELLOW}Config nicht gefunden: {fname}{NC}")
            continue
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception:
            continue
        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue
        df = load_ohlcv(symbol, timeframe, date_from, date_to)
        if df.empty:
            print(f"{YELLOW}Keine Daten fuer {symbol} ({timeframe}) — uebersprungen.{NC}")
            continue
        strategies_data[fname] = {
            'symbol': symbol, 'timeframe': timeframe,
            'df': df, 'config': config,
        }

    if not strategies_data:
        print(f"{RED}Keine Daten verfuegbar.{NC}")
        return

    print(f"\n--- Manuelle Portfolio-Simulation ---")
    print(f"Strategien: {list(strategies_data.keys())}")
    print(f"Zeitraum: {date_from} bis {date_to} | Startkapital: {capital:.0f} USDT\n")

    result = run_portfolio_simulation(capital, strategies_data, date_from, date_to)
    if result is None:
        print(f"{RED}Portfolio-Simulation fehlgeschlagen.{NC}")
        return

    _print_portfolio_result(result, capital)


# ---------------------------------------------------------------------------
# Modus 3: Automatische Portfolio-Optimierung (Greedy wie fibot)
# ---------------------------------------------------------------------------

def run_portfolio_finder(date_from: str, date_to: str, capital: float,
                          target_max_dd: float = 30.0, min_wr: float = 0.0,
                          auto: bool = False, configs_filter: list = None):
    from vbot.analysis.backtester import run_backtest, load_ohlcv
    from vbot.analysis.portfolio_simulator import run_portfolio_simulation

    if not os.path.isdir(CONFIGS_DIR):
        print(f"{RED}Kein Configs-Verzeichnis: {CONFIGS_DIR}{NC}")
        return

    cfg_files = sorted(f for f in os.listdir(CONFIGS_DIR)
                       if f.startswith('config_') and f.endswith('.json'))
    if configs_filter:
        cfg_files = [f for f in cfg_files if f in configs_filter]
    if not cfg_files:
        print(f"{YELLOW}Keine Configs gefunden. Erst run_pipeline.sh ausfuehren.{NC}")
        return

    print(f"\n--- vbot Portfolio-Finder ---")
    print(f"Zeitraum: {date_from} -> {date_to} | Kapital: {capital:.0f} USDT")
    print(f"Constraints: MaxDD <= {target_max_dd}% | MinWR >= {min_wr}%\n")

    # ── Schritt 1: Alle Configs isoliert backtesten ───────────────────────────
    all_results      = []
    data_cache       = {}
    backtest_results = {}   # fname -> BacktestResult (fuer Chart)

    for fname in cfg_files:
        cfg_path = os.path.join(CONFIGS_DIR, fname)
        try:
            with open(cfg_path) as f:
                config = json.load(f)
        except Exception:
            continue
        symbol    = config.get('market', {}).get('symbol', '')
        timeframe = config.get('market', {}).get('timeframe', '')
        if not symbol or not timeframe:
            continue

        print(f"  Backtest: {fname}...")
        df = load_ohlcv(symbol, timeframe, date_from, date_to)
        if df.empty or len(df) < 10:
            print(f"    {YELLOW}Keine Daten — uebersprungen.{NC}")
            continue

        data_cache[fname] = {'symbol': symbol, 'timeframe': timeframe,
                              'df': df, 'config': config}

        result = run_backtest(df, config, capital, symbol, timeframe)
        backtest_results[fname] = result
        coin = symbol.split('/')[0]
        all_results.append({
            'filename':     fname,
            'symbol':       symbol,
            'timeframe':    timeframe,
            'coin':         coin,
            'pnl_pct':      result.pnl_pct,
            'win_rate':     result.win_rate,
            'max_dd':       result.max_drawdown_pct,
            'avg_rr':       result.avg_rr,
            'total_trades': result.total_trades,
            'end_capital':  result.end_capital,
            'in_portfolio': False,
        })

    if not all_results:
        print(f"{RED}Keine Backtest-Ergebnisse.{NC}")
        _save_results([], [], date_from, date_to)
        return

    # ── Schritt 2: Kandidaten filtern ────────────────────────────────────────
    candidates = [r for r in all_results
                  if r['max_dd'] <= target_max_dd and r['win_rate'] >= min_wr and r['pnl_pct'] > 0]

    print(f"\n  {len(candidates)}/{len(all_results)} Kandidaten erfuellen Constraints "
          f"(DD <= {target_max_dd}%, WR >= {min_wr}%, PnL > 0%).")

    if not candidates:
        print(f"{YELLOW}Keine Kandidaten gefunden. Erhoehe max_dd oder senke min_wr.{NC}")
        _save_results(all_results, [], date_from, date_to)
        return

    # ── Schritt 3: Beste Einzelstrategie als Baseline ────────────────────────
    candidates_sorted = sorted(candidates, key=lambda x: x['end_capital'], reverse=True)

    best_single     = None
    best_single_sim = None
    print(f"\n  Suche beste Einzelstrategie als Baseline...")
    for r in candidates_sorted:
        sim = run_portfolio_simulation(
            capital, {r['filename']: data_cache[r['filename']]}, date_from, date_to
        )
        if sim is None:
            continue
        if sim['max_drawdown_pct'] <= target_max_dd:
            best_single     = r
            best_single_sim = sim
            break

    if best_single is None:
        print(f"{YELLOW}Keine Einzelstrategie besteht DD-Pruefung.{NC}")
        _save_results(all_results, [], date_from, date_to)
        return

    lbl = best_single['filename'].replace('config_', '').replace('_fibo.json', '')
    print(f"  Baseline: {lbl}  "
          f"EndKap={best_single_sim['end_capital']:.2f} USDT  "
          f"DD={best_single_sim['max_drawdown_pct']:.1f}%")

    # ── Schritt 4: Greedy Erweiterung — nur hinzufuegen wenn Verbesserung ────
    portfolio_files = [best_single['filename']]
    used_coins      = {best_single['coin']}
    best_end_cap    = best_single_sim['end_capital']
    final_sim       = best_single_sim
    best_single['in_portfolio'] = True

    remaining = [r for r in candidates_sorted
                 if r['filename'] != best_single['filename']]

    improved = True
    while improved:
        improved = False
        for r in remaining:
            if r['coin'] in used_coins or r['filename'] in portfolio_files:
                continue

            test_files = portfolio_files + [r['filename']]
            test_data  = {fn: data_cache[fn] for fn in test_files if fn in data_cache}
            sim = run_portfolio_simulation(capital, test_data, date_from, date_to)
            if sim is None:
                continue

            # Nur aufnehmen wenn: Endkapital steigt UND DD bleibt im Limit
            if sim['end_capital'] > best_end_cap and sim['max_drawdown_pct'] <= target_max_dd:
                portfolio_files.append(r['filename'])
                used_coins.add(r['coin'])
                r['in_portfolio'] = True
                best_end_cap = sim['end_capital']
                final_sim    = sim
                lbl = r['filename'].replace('config_', '').replace('_fibo.json', '')
                print(f"  + {lbl}  "
                      f"EndKap={sim['end_capital']:.2f} USDT  "
                      f"DD={sim['max_drawdown_pct']:.1f}%")
                improved = True
                break   # Neustart der Schleife nach jeder Hinzufuegung

    # ── Ergebnis ausgeben ─────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"{'OPTIMALES PORTFOLIO':^60}")
    print(f"{'='*60}")
    for fn in portfolio_files:
        r = next(x for x in all_results if x['filename'] == fn)
        strat = f"{r['symbol'].split('/')[0]}/{r['timeframe']}"
        print(f"  {strat:<20}  PnL={r['pnl_pct']:+.2f}%  WR={r['win_rate']:.1f}%  DD={r['max_dd']:.1f}%")
    if len(portfolio_files) == 1:
        print(f"\n  (Einzelstrategie — kein weiterer Coin verbessert das Portfolio)")

    print(f"\n  Portfolio-Gesamt:")
    _print_portfolio_result(final_sim, capital)

    _save_results(all_results, portfolio_files, date_from, date_to)

    # ── Charts + Excel senden ─────────────────────────────────────────────────
    if auto:
        _generate_portfolio_chart(final_sim, all_results, portfolio_files,
                                  capital, date_from, date_to, backtest_results)
        _generate_trades_excel(final_sim, portfolio_files, capital)
    else:
        print()
        ans = input("  Charts & Excel erstellen und via Telegram senden? "
                    "(j/n) [Standard: n]: ").strip().lower()
        if ans in ('j', 'y', 'ja'):
            _generate_portfolio_chart(final_sim, all_results, portfolio_files,
                                      capital, date_from, date_to, backtest_results)
            _generate_trades_excel(final_sim, portfolio_files, capital)


def _generate_portfolio_chart(final_sim: dict, all_results: list, portfolio_files: list,
                               capital: float, date_from: str, date_to: str,
                               backtest_results: dict = None):
    """Erstellt interaktiven Portfolio-Equity-Chart im mbot-Stil und sendet via Telegram."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {YELLOW}plotly nicht installiert — Chart uebersprungen. (pip install plotly){NC}")
        return

    equity_curve  = final_sim.get('equity_curve', [])
    trade_history = final_sim.get('trade_history', [])
    if not equity_curve:
        print(f"  {YELLOW}Keine Equity-Daten — Chart uebersprungen.{NC}")
        return

    portfolio_set = set(portfolio_files)
    eq_map = {e['timestamp']: e['equity'] for e in equity_curve}

    # Portfolio-Equity-Kurve
    port_times  = [str(e['timestamp'])[:19] for e in equity_curve]
    port_equity = [e['equity']              for e in equity_curve]

    # Strat-Labels fuer Titel
    strat_labels = []
    for fn in portfolio_files:
        r = next((x for x in all_results if x['filename'] == fn), None)
        if r:
            strat_labels.append(f"{r['symbol'].split('/')[0]}/{r['timeframe']}")

    n_strats  = len(portfolio_files)
    pnl_pct   = final_sim['total_pnl_pct']
    max_dd    = final_sim['max_drawdown_pct']
    end_cap   = final_sim['end_capital']
    n_trades  = final_sim['trade_count']
    wr        = final_sim['win_rate']
    sign      = '+' if pnl_pct >= 0 else ''
    pairs_str = ', '.join(strat_labels)

    title = (
        f"vbot Portfolio \u2014 {n_strats} Strategie(n) ({pairs_str}) | "
        f"Trades: {n_trades} | WR: {wr:.1f}% | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Endkapital: {end_cap:.2f} USDT | MaxDD: {max_dd:.1f}%"
    )

    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        specs=[[{'secondary_y': True}], [{'secondary_y': False}]],
        vertical_spacing=0.03,
        row_heights=[0.85, 0.15],
    )

    # ── Einzel-Equity-Kurven aus isolierten Backtests (linke Y-Achse) ─────────
    STRAT_COLORS = [
        '#ffa726', '#e91e63', '#ab47bc', '#26c6da',
        '#66bb6a', '#ff7043', '#42a5f5', '#d4e157',
    ]
    if backtest_results:
        for idx, (fn, bt_result) in enumerate(sorted(backtest_results.items())):
            closed = [t for t in bt_result.trades if t.result != 'open']
            if not closed:
                continue
            eq   = capital
            xs   = [str(closed[0].timestamp)[:19]]
            ys   = [capital]
            for t in closed:
                eq += t.pnl_usdt
                xs.append(str(t.timestamp)[:19])
                ys.append(round(eq, 4))
            r     = next((x for x in all_results if x['filename'] == fn), None)
            label = f"{r['symbol'].split('/')[0]}/{r['timeframe']}" if r else fn
            color = STRAT_COLORS[idx % len(STRAT_COLORS)]
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines',
                name=label,
                line=dict(color=color, width=1.2),
                hovertemplate=f'{label}<br>Equity: %{{y:.2f}} USDT<extra></extra>',
            ), row=1, col=1, secondary_y=False)

    # ── Portfolio-Equity (rechte Y-Achse, blau) ───────────────────────────────
    fig.add_trace(go.Scatter(
        x=port_times, y=port_equity,
        mode='lines', name='Portfolio Equity',
        line=dict(color='#5c9bd6', width=2),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ), row=1, col=1, secondary_y=True)

    # ── Trade-Marker auf Portfolio-Equity (rechte Y-Achse) ────────────────────
    entry_x, entry_y = [], []
    tp_x, tp_y, tp_pnl = [], [], []
    sl_x, sl_y, sl_pnl = [], [], []

    for t in trade_history:
        ts_open  = str(t['ts'])[:19]
        ts_close = str(t.get('ts_close', t['ts']))[:19]
        eq_entry = eq_map.get(t['ts'], capital)
        eq_exit  = eq_map.get(t.get('ts_close', t['ts']), capital)
        entry_x.append(ts_open)
        entry_y.append(eq_entry)
        if t['pnl'] > 0:
            tp_x.append(ts_close); tp_y.append(eq_exit)
            tp_pnl.append(round(t['pnl'], 4))
        else:
            sl_x.append(ts_close); sl_y.append(eq_exit)
            sl_pnl.append(round(t['pnl'], 4))

    if entry_x:
        fig.add_trace(go.Scatter(
            x=entry_x, y=entry_y, mode='markers',
            marker=dict(symbol='triangle-up', size=12, color='#26a69a',
                        line=dict(color='#ffffff', width=0.5)),
            name='Entry \u25b2',
            hovertemplate='Entry<br>%{x}<extra></extra>',
        ), row=1, col=1, secondary_y=True)

    if tp_x:
        fig.add_trace(go.Scatter(
            x=tp_x, y=tp_y, mode='markers',
            marker=dict(symbol='circle', size=11, color='#00bcd4',
                        line=dict(color='#ffffff', width=0.5)),
            name='Exit TP \u2713',
            customdata=tp_pnl,
            hovertemplate='Exit TP<br>%{x}<br>PnL: %{customdata:.4f} USDT<extra></extra>',
        ), row=1, col=1, secondary_y=True)

    if sl_x:
        fig.add_trace(go.Scatter(
            x=sl_x, y=sl_y, mode='markers',
            marker=dict(symbol='x', size=13, color='#ef5350',
                        line=dict(color='#ef5350', width=2.5)),
            name='Exit SL \u2717',
            customdata=sl_pnl,
            hovertemplate='Exit SL<br>%{x}<br>PnL: %{customdata:.4f} USDT<extra></extra>',
        ), row=1, col=1, secondary_y=True)

    # ── Start-Annotation ──────────────────────────────────────────────────────
    fig.add_annotation(
        x=port_times[0] if port_times else date_from,
        y=capital,
        text=f'Start {capital:.0f} USDT',
        showarrow=False, font=dict(size=10, color='#aaaaaa'),
        xanchor='left', yanchor='bottom',
        xref='x', yref='y2',
    )

    # ── Trade-Timeline (unteres Panel) ────────────────────────────────────────
    if entry_x:
        fig.add_trace(go.Scatter(
            x=entry_x, y=[1] * len(entry_x), mode='markers',
            marker=dict(symbol='triangle-up', size=9, color='#26a69a'),
            showlegend=False,
            hovertemplate='Entry<br>%{x}<extra></extra>',
        ), row=2, col=1)
    if tp_x:
        fig.add_trace(go.Scatter(
            x=tp_x, y=[1] * len(tp_x), mode='markers',
            marker=dict(symbol='circle', size=8, color='#00bcd4'),
            showlegend=False,
            hovertemplate='Exit TP<br>%{x}<extra></extra>',
        ), row=2, col=1)
    if sl_x:
        fig.add_trace(go.Scatter(
            x=sl_x, y=[1] * len(sl_x), mode='markers',
            marker=dict(symbol='x', size=9, color='#ef5350',
                        line=dict(width=2)),
            showlegend=False,
            hovertemplate='Exit SL<br>%{x}<extra></extra>',
        ), row=2, col=1)

    # ── Layout ────────────────────────────────────────────────────────────────
    fig.update_layout(
        title=dict(text=title, font=dict(size=11), x=0.5, xanchor='center'),
        template='plotly_dark',
        xaxis_rangeslider_visible=False,
        xaxis2_rangeslider_visible=False,
        legend=dict(orientation='h', yanchor='bottom', y=1.01,
                    xanchor='center', x=0.5, font=dict(size=10)),
        height=720,
        margin=dict(l=70, r=80, t=80, b=40),
        yaxis=dict(title='Einzel-Equity (USDT)'),
        yaxis2=dict(title='Portfolio-Equity (USDT)', showgrid=False,
                    tickfont=dict(color='#5c9bd6'),
                    title_font=dict(color='#5c9bd6')),
        yaxis3=dict(visible=False),
    )
    fig.update_yaxes(visible=False, row=2, col=1)

    os.makedirs(os.path.join(PROJECT_ROOT, 'artifacts', 'charts'), exist_ok=True)
    out_file = os.path.join(PROJECT_ROOT, 'artifacts', 'charts', 'vbot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  {GREEN}Chart gespeichert: {out_file}{NC}")

    # ── Telegram ──────────────────────────────────────────────────────────────
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        tg        = secrets.get('telegram', {})
        bot_token = tg.get('bot_token', '')
        chat_id   = tg.get('chat_id',   '')
    except Exception:
        bot_token = chat_id = ''

    if bot_token and chat_id:
        from vbot.utils.telegram import send_document
        caption = (
            f"vbot Portfolio Chart | {n_strats} Strategie(n) | "
            f"{date_from} \u2192 {date_to} | "
            f"PnL: {sign}{pnl_pct:.1f}% | MaxDD: {max_dd:.1f}% | "
            f"Endkap: {end_cap:.2f} USDT"
        )
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Chart via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — Chart nur lokal gespeichert.{NC}")


def _generate_trades_excel(final_sim: dict, portfolio_files: list, capital: float):
    """Erstellt eine Excel-Tabelle mit allen Portfolio-Trades und sendet sie via Telegram."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {YELLOW}openpyxl nicht installiert — Excel uebersprungen. (pip install openpyxl){NC}")
        return

    trade_history = final_sim.get('trade_history', [])
    if not trade_history:
        print(f"  {YELLOW}Keine Trades — Excel uebersprungen.{NC}")
        return

    # Zeilen aufbauen
    equity = capital
    rows   = []
    for i, t in enumerate(trade_history):
        pnl      = float(t['pnl'])
        equity  += pnl
        fname    = t.get('fname', '')
        strat    = fname.replace('config_', '').replace('_fibo.json', '')
        dir_     = t.get('direction', '').upper()
        entry    = round(float(t.get('entry',      0)), 6)
        exit_p   = round(float(t.get('exit',       0)), 6)
        fibo_lvl = t.get('fibo_level', '')
        ergebnis = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        rows.append({
            'Nr':            i + 1,
            'Datum':         str(t.get('ts', ''))[:16].replace('T', ' '),
            'Strategie':     strat,
            'Richtung':      dir_,
            'Fibo-Level':    fibo_lvl,
            'Entry':         entry,
            'Exit':          exit_p,
            'Ergebnis':      ergebnis,
            'PnL (USDT)':    round(pnl,    4),
            'Kapital':       round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill    = PatternFill('solid', fgColor='D6F4DC')
    loss_fill   = PatternFill('solid', fgColor='FAD7D7')
    alt_fill    = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Strategie': 26, 'Richtung': 10, 'Fibo-Level': 12,
        'Entry': 14, 'Exit': 14, 'Ergebnis': 14, 'PnL (USDT)': 14, 'Kapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        if row['Ergebnis'] == 'TP erreicht':
            fill = win_fill
        elif r_idx % 2 == 0:
            fill = loss_fill
        else:
            fill = alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Kapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    # Zusammenfassung
    total  = len(rows)
    wins   = sum(1 for r in rows if r['Ergebnis'] == 'TP erreicht')
    sr     = total + 3
    pnl_total = rows[-1]['Kapital'] - capital if rows else 0.0
    pnl_pct   = pnl_total / capital * 100 if capital else 0.0
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', total),
        ('Win-Rate',      f"{wins / total * 100:.1f}%" if total else '—'),
        ('PnL',           f"{pnl_pct:+.1f}%"),
        ('Endkapital',    f"{rows[-1]['Kapital']:.2f} USDT" if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    os.makedirs(os.path.join(PROJECT_ROOT, 'artifacts', 'charts'), exist_ok=True)
    out_file = os.path.join(PROJECT_ROOT, 'artifacts', 'charts', 'vbot_trades.xlsx')
    wb.save(out_file)
    print(f"  {GREEN}Excel gespeichert: {out_file}{NC}")

    # Telegram
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    try:
        with open(secret_path) as f:
            secrets = json.load(f)
        tg        = secrets.get('telegram', {})
        bot_token = tg.get('bot_token', '')
        chat_id   = tg.get('chat_id',   '')
    except Exception:
        bot_token = chat_id = ''

    if bot_token and chat_id:
        from vbot.utils.telegram import send_document
        caption = (f"vbot Trades — {total} Trades | "
                   f"WR: {wins / total * 100:.1f}% | PnL: {pnl_pct:+.1f}%" if total else "vbot Trades")
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — nur lokal gespeichert.{NC}")


def _print_portfolio_result(result: dict, start_capital: float):
    sign  = '+' if result['total_pnl_pct'] >= 0 else ''
    color = GREEN if result['total_pnl_pct'] >= 0 else RED
    print(f"  PnL:          {color}{sign}{result['total_pnl_pct']:.2f}%{NC}")
    print(f"  Max Drawdown: {result['max_drawdown_pct']:.2f}%")
    print(f"  Trades:       {result['trade_count']} (W:{result['wins']} L:{result['losses']})")
    print(f"  Win Rate:     {result['win_rate']:.1f}%")
    print(f"  Endkapital:   {result['end_capital']:.2f} USDT (Start: {start_capital:.0f} USDT)")


def _save_results(all_results: list, portfolio_files: list, date_from: str, date_to: str):
    os.makedirs(RESULTS_DIR, exist_ok=True)
    out = {
        'date_from':         date_from,
        'date_to':           date_to,
        'optimal_portfolio': portfolio_files,
        'all_results':       all_results,
    }
    path = os.path.join(RESULTS_DIR, 'optimization_results.json')
    with open(path, 'w') as f:
        json.dump(out, f, indent=2)
    print(f"\n  Ergebnisse gespeichert: {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='vbot — Ergebnisanzeige')
    parser.add_argument('--mode',          type=int,   default=1, choices=[1, 2, 3, 4])
    parser.add_argument('--capital',       type=float, default=1000.0)
    parser.add_argument('--from',  dest='date_from',   default=None)
    parser.add_argument('--to',    dest='date_to',     default=None)
    parser.add_argument('--target-max-dd', type=float, default=30.0)
    parser.add_argument('--min-wr',        type=float, default=0.0)
    parser.add_argument('--auto',          action='store_true')
    parser.add_argument('--configs',       type=str,   default=None,
                        help="Leerzeichen-getrennte Liste von Config-Dateinamen")
    args = parser.parse_args()

    today    = date.today().isoformat()
    d_from   = args.date_from if args.date_from else '2024-01-01'
    d_to     = args.date_to   if args.date_to   else today
    cfg_list = args.configs.split() if args.configs else None

    if args.mode == 1:
        run_all_configs_isolated(d_from, d_to, args.capital, cfg_list)
    elif args.mode == 2:
        if not cfg_list:
            print(f"{RED}--configs benoetigt fuer Modus 2{NC}")
            sys.exit(1)
        run_manual_portfolio(d_from, d_to, args.capital, cfg_list)
    elif args.mode == 3:
        run_portfolio_finder(
            d_from, d_to, args.capital,
            target_max_dd=args.target_max_dd,
            min_wr=args.min_wr,
            auto=args.auto,
            configs_filter=cfg_list,
        )
    elif args.mode == 4:
        from vbot.analysis.interactive_chart import run_interactive_chart
        secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
        secrets = {}
        if os.path.exists(secret_path):
            with open(secret_path) as f:
                secrets = json.load(f)
        run_interactive_chart(secrets)
